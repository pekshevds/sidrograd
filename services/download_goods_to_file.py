from typing import Any
from openpyxl import Workbook
from repositories import good_repository
import io


def fetch_goods_to_xlsx_data() -> Any:
    wb = Workbook()
    ws = wb.active
    ws.title = "goods"

    # Заголовок таблицы
    ws.cell(row=1, column=1, value="Наименование")
    ws.cell(row=1, column=2, value="Артикул")
    ws.cell(row=1, column=3, value="Активный")
    ws.cell(row=1, column=4, value="Новинка")
    ws.cell(row=1, column=5, value="Пометка")
    ws.cell(row=1, column=6, value="Наименование полное")
    ws.cell(row=1, column=7, value="Объем, л")
    ws.cell(row=1, column=8, value="Крепость, %")
    ws.cell(row=1, column=9, value="В упаковке, шт")
    ws.cell(row=1, column=10, value="Срок годности, мес")
    ws.cell(row=1, column=11, value="Раздел каталога")
    ws.cell(row=1, column=12, value="Торговая марка")
    ws.cell(row=1, column=13, value="Газация")
    ws.cell(row=1, column=14, value="Пастеризация")
    ws.cell(row=1, column=15, value="Фильтрация")
    ws.cell(row=1, column=16, value="Производитель")
    ws.cell(row=1, column=17, value="Единица измерения")
    ws.cell(row=1, column=18, value="Тип ферментации")
    ws.cell(row=1, column=19, value="Стиль")
    ws.cell(row=1, column=20, value="Страна")
    ws.cell(row=1, column=21, value="Описание")

    # Данные таблицы
    for i, good in enumerate(good_repository.fetch_all_goods(), 2):
        ws.cell(row=i, column=1, value=good.name)
        ws.cell(row=i, column=2, value=good.art)
        ws.cell(row=i, column=3, value=good.is_active)
        ws.cell(row=i, column=4, value=good.is_new)
        ws.cell(row=i, column=5, value=good.marked)
        ws.cell(row=i, column=6, value=good.full_name)
        ws.cell(row=i, column=7, value=str(good.volume))
        ws.cell(row=i, column=8, value=str(good.strength))
        ws.cell(row=i, column=9, value=str(good.in_package))
        ws.cell(row=i, column=10, value=str(good.expiration_date))
        ws.cell(row=i, column=11, value=str(good.category))
        ws.cell(row=i, column=12, value=str(good.trade_mark))
        ws.cell(row=i, column=13, value=str(good.gassing))
        ws.cell(row=i, column=14, value=str(good.pasteurization))
        ws.cell(row=i, column=15, value=str(good.filtering))
        ws.cell(row=i, column=16, value=str(good.manufacturer))
        ws.cell(row=i, column=17, value=str(good.unit))
        ws.cell(row=i, column=18, value=str(good.type_of_fermentation))
        ws.cell(row=i, column=19, value=str(good.style))
        ws.cell(row=i, column=20, value=str(good.country))
        ws.cell(row=i, column=21, value=str(good.description))

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()
